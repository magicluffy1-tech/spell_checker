"""
sheets_handler.py

데이터 입출력 핸들러 (xlsx 바이너리 다운로드 기반)
─────────────────────────────────────────────────────
★ 핵심 변경사항 (HTML/CSV 파싱 방식 완전 폐기):
  - 구글 시트를 .xlsx 형식으로 직접 export 하여 메모리에 로드
  - openpyxl.sheetnames로 탭 목록 100% 정확 추출 (JS 렌더링 불필요)
  - pd.read_excel()로 한글 인코딩 깨짐 원천 차단
  - 공개 시트(링크가 있는 모든 사용자)면 인증 없이 동작

지원 방식:
  1) 엑셀(.xlsx) / CSV 파일 업로드 (Streamlit UploadedFile 객체)
  2) 구글 시트 공유 링크 (뷰어 권한, 읽기 전용) → xlsx export URL 자동 변환
     → 탭(시트) 이름 드롭다운 자동 제공 + 한글 데이터 무결 보장
결과 DataFrame → 엑셀/CSV 바이트 다운로드 유틸리티 포함
"""

import io
import re
from typing import Optional
import pandas as pd
import requests
import openpyxl

# ──────────────────────────────────────────────
# 내부: 구글 시트 ID 추출 및 URL 변환
# ──────────────────────────────────────────────

def _extract_sheet_id(url: str) -> Optional[str]:
    """구글 시트 공유 URL에서 스프레드시트 ID를 추출한다."""
    cleaned = url.strip()
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", cleaned)
    if m:
        return m.group(1)
    m = re.search(r"/d/([a-zA-Z0-9_-]{40,})", cleaned)
    if m:
        return m.group(1)
    return None


def _make_xlsx_export_url(sheet_id: str) -> str:
    """스프레드시트 ID → xlsx 내보내기 URL 생성."""
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def _download_xlsx(url: str) -> bytes:
    """
    구글 시트를 xlsx 바이너리로 다운로드한다.

    Args:
        url: 구글 시트 공유 URL (edit / view / 직접 export URL 모두 허용)

    Returns:
        xlsx 바이너리 (bytes)

    Raises:
        ValueError: 유효하지 않은 URL
        PermissionError: 비공개 시트 또는 접근 차단
        ConnectionError: 네트워크 오류
    """
    sheet_id = _extract_sheet_id(url)
    if not sheet_id:
        raise ValueError("유효한 구글 시트 URL이 아닙니다. 공유 링크를 다시 확인하세요.")

    export_url = _make_xlsx_export_url(sheet_id)

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        )
    }

    # 1차 시도 (verify=True)
    try:
        resp = requests.get(export_url, headers=headers, timeout=20, verify=True)
        if resp.status_code in (401, 403):
            raise PermissionError(
                "구글 시트에 접근할 수 없습니다.\n\n"
                "💡 **해결 방법:**\n"
                "1. 구글 시트 → **[공유]** 버튼 클릭\n"
                "2. 일반 액세스를 **'링크가 있는 모든 사용자 (뷰어)'** 로 변경\n"
                "3. 변경 후 링크를 다시 복사해서 붙여넣어 주세요."
            )
        resp.raise_for_status()
        # xlsx 파일인지 확인 (구글 로그인 리다이렉트 HTML이 오는 경우 차단)
        content_type = resp.headers.get("Content-Type", "")
        if "html" in content_type.lower() or len(resp.content) < 1000:
            raise PermissionError(
                "구글 시트가 비공개 상태이거나 로그인이 필요합니다.\n\n"
                "💡 공유 설정에서 '링크가 있는 모든 사용자'로 변경해 주세요."
            )
        return resp.content
    except PermissionError:
        raise
    except requests.exceptions.SSLError:
        # 2차 시도: SSL 인증서 문제 폴백
        try:
            resp = requests.get(export_url, headers=headers, timeout=20, verify=False)
            resp.raise_for_status()
            return resp.content
        except Exception as e:
            raise ConnectionError(f"네트워크 오류 (SSL 폴백 실패): {e}") from e
    except requests.exceptions.RequestException as e:
        raise ConnectionError(f"네트워크 오류: {e}") from e


# ──────────────────────────────────────────────
# 1. 탭(시트) 목록 조회 — xlsx 기반 100% 정확
# ──────────────────────────────────────────────

def get_sheet_tabs(url: str) -> list[dict]:
    """
    구글 시트 공유 링크에서 모든 탭(시트) 이름 목록을 조회한다.

    HTML 파싱·JavaScript 렌더링 없이, xlsx export 바이너리에서
    openpyxl로 직접 추출하므로 한글 이름도 완벽히 지원한다.

    Returns:
        [{"name": "시트이름", "index": 0}, ...]
        조회 실패 시 빈 리스트 반환
    """
    try:
        content = _download_xlsx(url)
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        tabs = [{"name": name, "index": i} for i, name in enumerate(wb.sheetnames)]
        wb.close()
        return tabs
    except Exception:
        return []


# ──────────────────────────────────────────────
# 2. 데이터 로드 함수
# ──────────────────────────────────────────────

def get_raw_data_from_dataframe(df: pd.DataFrame, text_column: str) -> list[dict]:
    """
    이미 로드된 DataFrame에서 검사 대상 텍스트 목록을 추출한다.

    Args:
        df: 업로드된 파일에서 읽어온 DataFrame
        text_column: 맞춤법 검사를 수행할 열 이름

    Returns:
        [{"row_index": int, "original_text": str}, ...]
    """
    if text_column not in df.columns:
        raise ValueError(f"'{text_column}' 열을 찾을 수 없습니다. 실제 열 이름을 확인하세요.")

    records = []
    for idx, row in df.iterrows():
        cell_value = row[text_column]
        if pd.isna(cell_value) or str(cell_value).strip() == "":
            continue
        records.append({
            "row_index": idx,
            "original_text": str(cell_value).strip(),
        })
    return records


def get_raw_data_from_public_url(
    url: str,
    text_column: Optional[str] = None,
    sheet_name: Optional[str] = None,
    gid: Optional[str] = None,  # 하위 호환 유지 (미사용)
) -> tuple[pd.DataFrame, list[dict]]:
    """
    구글 시트 공유 링크(뷰어 권한 이상)에서 데이터를 로드한다.

    xlsx export URL로 전체 워크북을 다운로드하여 메모리에서 파싱하므로
    한글 인코딩 깨짐이 없고, 탭 이름 지정도 정확히 작동한다.

    Args:
        url: 구글 시트 공유 URL
        text_column: 검사할 열 이름. None이면 첫 번째 열 사용.
        sheet_name: 읽을 탭 이름. None이면 첫 번째 탭 사용.
        gid: 미사용 (하위 호환용 파라미터, 무시됨)

    Returns:
        (DataFrame 전체, records 리스트)
    """
    content = _download_xlsx(url)

    # 탭 이름 지정이 있으면 해당 탭, 없으면 0번째 탭
    try:
        if sheet_name:
            df = pd.read_excel(
                io.BytesIO(content),
                sheet_name=sheet_name,
                engine="openpyxl",
                dtype=str,
            )
        else:
            df = pd.read_excel(
                io.BytesIO(content),
                sheet_name=0,
                engine="openpyxl",
                dtype=str,
            )
    except Exception as e:
        # 탭 이름이 틀렸을 경우 첫 번째 탭으로 폴백
        try:
            df = pd.read_excel(
                io.BytesIO(content),
                sheet_name=0,
                engine="openpyxl",
                dtype=str,
            )
        except Exception as fallback_e:
            raise ValueError(f"시트 데이터를 읽는 중 오류가 발생했습니다: {fallback_e}") from fallback_e

    if df.empty:
        raise ValueError("선택한 탭(시트)이 비어 있습니다.")

    col = text_column if text_column and text_column in df.columns else df.columns[0]
    records = get_raw_data_from_dataframe(df, col)
    return df, records


def load_uploaded_file(
    uploaded_file,
    text_column: Optional[str] = None,
) -> tuple[pd.DataFrame, list[dict]]:
    """
    Streamlit UploadedFile 객체(.xlsx, .xls, .csv)를 DataFrame으로 변환하고
    검사 대상 records를 추출한다.
    """
    filename = uploaded_file.name.lower()
    if filename.endswith(".csv"):
        df = pd.read_csv(uploaded_file, encoding="utf-8-sig", dtype=str)
    elif filename.endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded_file, engine="openpyxl", dtype=str)
    else:
        raise ValueError("지원하지 않는 파일 형식입니다. .xlsx, .xls, .csv 파일을 업로드하세요.")

    if df.empty:
        raise ValueError("파일이 비어 있습니다.")

    col = text_column if text_column and text_column in df.columns else df.columns[0]
    records = get_raw_data_from_dataframe(df, col)
    return df, records


# ──────────────────────────────────────────────
# 3. 결과 다운로드 유틸리티
# ──────────────────────────────────────────────

def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """DataFrame을 .xlsx 바이트로 변환 (st.download_button용)."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="맞춤법검사결과")
    return buffer.getvalue()


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """DataFrame을 UTF-8 BOM CSV 바이트로 변환 (한글 엑셀 호환)."""
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
